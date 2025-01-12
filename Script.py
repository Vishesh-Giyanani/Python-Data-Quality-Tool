import os
import re
import json
import warnings
import numpy as np
import pandas as pd
from io import BytesIO
from decimal import Decimal
from openpyxl import Workbook
from google.cloud import storage
from google.cloud import bigquery
from openpyxl import load_workbook
from datetime import datetime, timezone
from google.oauth2 import service_account

def execute(run):

    key = {
#Enter GCP Key 1
    }
    devkey = {
#Enter GCP Key 2
    }


    def main(bucket_name, retailers_dict, code, countryname, view1, view2, input_file_country, output_file_country, columnsdict, date1, pushproject, pushdataset):
        if code == run:
            warnings.filterwarnings("ignore")
            excelmode = False

            def read_json_from_gcs(bucket_name, file_path):
                credentials = service_account.Credentials.from_service_account_info(devkey)
                client = storage.Client(credentials=credentials,project=devkey['project_id'])
                bucket = client.bucket(bucket_name)
                blob = bucket.blob(file_path)
                json_data = json.loads(blob.download_as_string())
                print(f"{file_path} read succesfully")
                return json_data

            def write_json_to_gcs(data, bucket_name, file_path):
                credentials = service_account.Credentials.from_service_account_info(devkey)
                client = storage.Client(credentials=credentials,project=devkey['project_id'])
                bucket = client.bucket(bucket_name)
                blob = bucket.blob(file_path)
                blob.upload_from_string(json.dumps(data, indent=4))
                print(f"{file_path} has been successfully written to GCS.")

            def read_excel_from_gcs(bucket_name, file_path, sheet_name=0):
                credentials = service_account.Credentials.from_service_account_info(devkey)
                client = storage.Client(credentials=credentials,project=devkey['project_id'])
                bucket = client.bucket(bucket_name)
                blob = bucket.blob(file_path)
                data = blob.download_as_bytes()  # Read the file into memory
                df = pd.read_excel(data, sheet_name=sheet_name)
                print(f"{file_path} read succesfully")
                return df

            def write_excel_to_gcs(dataframe_dict, bucket_name, file_path):
                credentials = service_account.Credentials.from_service_account_info(devkey)
                client = storage.Client(credentials=credentials,project=devkey['project_id'])
                bucket = client.bucket(bucket_name)
                blob = bucket.blob(file_path)
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    for sheet_name, df in dataframe_dict.items():
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                output.seek(0)
                blob.upload_from_file(
                    output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

            credentials = service_account.Credentials.from_service_account_info(key)
            client = bigquery.Client(credentials=credentials, project=key['project_id'])

            highest_date_query = f"""
                SELECT MAX(max_date) 
                FROM (
                    SELECT MAX(sales_date) AS max_date FROM {view1}
                    UNION ALL
                    SELECT MAX(sales_date) FROM {view2}
                ) AS combined_dates;
                """

            max_date_result = client.query(highest_date_query).result()
            max_date = list(max_date_result)[0][0]
            date_limit = f""" WHERE sales_date BETWEEN '{date1}' AND '{max_date}'"""  ##TEST##

            print("Imports loaded ✅")

            # os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = r'/tmp/devkey.json'
            df = read_excel_from_gcs(bucket_name, input_file_country, sheet_name=4)
            print("Input File read ✅")

            def gettabledetails(data, columname):
                text = data.at[0, columname].lower()
                if text != 'no':
                    return data.at[0, columname]
                else:
                    return "NO"

            def getsection1(data, columname):
                text = str(data.at[0, columname]).lower()
                if text != 0 and text == 'yes':
                    return 'yes'
                else:
                    return "no"

            def getsection3(data, columname):
                text = str(data.at[0, columname]).lower()
                if text != 0 and text == 'yes':
                    return 'yes'
                else:
                    return "no"

            def getsection2(data, columname):
                lst = []
                output = []
                for a in range(len(data[columname])):
                    if str(data[columname][a]).lower() == 'yes':
                        lst.append(a)
                output = [data['Index'][a] for a in lst]
                return output

            all_variables = {
                "project_id":                   gettabledetails(df, 'Project ID *'),
                "dataset_id":                   gettabledetails(df, 'Dataset ID *'),
                "table_id":                     gettabledetails(df, 'Table ID *'),
                "Check_All_Columns_Exist":      getsection1(df, 'Check All Columns Exist'),
                "Get_Empty_Rows":               getsection1(df, 'Get Empty Rows'),
                "Get_Total_Rows":               getsection1(df, 'Get Total Rows'),
                "Get_Total_Columns":            getsection1(df, 'Get Total Columns'),
                "Get_Total_Sales":              getsection1(df, 'Get Total Sales'),
                "Get_Total_Quantity":           getsection1(df, 'Get Total Quantity'),
                "Check_Last_Refresh":           getsection1(df, 'Check Last Refresh'),
                "Null_Check":                   getsection2(df, 'Null Check'),
                "NaN_Check":                    getsection2(df, 'NaN Check'),
                "Duplicate_Check":              getsection2(df, 'Duplicate Check'),
                "Bad_Character_Check":          getsection2(df, 'Bad Character Check'),
                "Check_Negatives":              getsection2(df, 'Check Negatives'),
                "Sales_Mismatch":               getsection2(df, 'Sales Mismatch'),
                "Check_Odd_Data_Types":         getsection2(df, 'Check Odd Data Types'),
                "Country_Code_Check":           getsection2(df, 'Country Code Check'),
                "Invalid_Date_Check":           getsection2(df, 'Invalid Date Check'),
                "Numeric_Check":                getsection2(df, 'Numeric Check'),
                "Check_Product_Key":            getsection3(df, 'Check Product Key'),
                "Check_Store_Key":              getsection3(df, 'Check Store Key'),
                "Check_Retailer_Key":           getsection3(df, 'Check Retailer Key'),
                "Check_Missing_Retailer":       getsection3(df, 'Check Missing Retailer'),
                "Check_Missing_Product":        getsection3(df, 'Check Missing Product'),
                "Check_Month_on_Month_Growth":  getsection3(df, 'Check Month on Month Growth'),
                "Check_Dormant_Store":          getsection3(df, 'Check Dormant Store'),
                "Check_Dormant_Product":        getsection3(df, 'Check Dormant Product')
            }

            print("Input File Processed ✅")

            def extractjson(keys):
                return {key: all_variables[key] for key in keys}

            project_details = extractjson(keys=["project_id", "dataset_id", "table_id"])
            section1 = extractjson(keys=["Check_All_Columns_Exist", "Get_Empty_Rows", "Get_Total_Rows","Get_Total_Columns", "Get_Total_Sales", "Get_Total_Quantity", "Check_Last_Refresh"])
            section2a = extractjson(keys=["Null_Check", "NaN_Check", "Duplicate_Check", "Bad_Character_Check","Check_Negatives", "Sales_Mismatch", "Check_Odd_Data_Types", "Numeric_Check"])
            section2b = extractjson(keys=["Country_Code_Check", "Invalid_Date_Check"])
            section3 = extractjson(keys=["Check_Product_Key", "Check_Store_Key", "Check_Retailer_Key", "Check_Missing_Retailer","Check_Missing_Product",  "Check_Month_on_Month_Growth", "Check_Dormant_Store", "Check_Dormant_Product"])

            print("Begining Data Fetch from Bigquery ✅")

            query = f"""
            WITH Union_Transactions AS ( SELECT * FROM {view1} {date_limit}
            UNION ALL
            SELECT * FROM {view2} {date_limit})
            SELECT * FROM Union_Transactions;
            """

            credentials = service_account.Credentials.from_service_account_info(key)
            client = bigquery.Client(credentials=credentials, project=key['project_id'])
            query_job = client.query(query)
            print("Query created for Bigquery ✅")
            df = query_job.to_dataframe()

            print("Finisbhed Data Fetch from Bigquery ✅")
            print("EPOS Data Quality Analysis has begun. ✅")

            # Section 1 - complete
            print("Loading Section 1 Functions ✅")

            # 1. Check Missing Columns
            def columncheck(data, columns_dict):
                columns_to_check = set(columns_dict.keys())
                data_columns = set(data.columns)
                missing_columns = columns_to_check - data_columns

                return list(missing_columns)

            # 2. Check Empty Rows
            # Get Back ##TEST##

            # 3, 4. Get Shape

            def getshape(data, num):
                count = data.shape
                return count[num]

            # 5, 6. Get Total Sales & Quantity
            def gettotal(data, column):
                return data[column].sum()

            # 7. Check Last Refresh
            def get_table_last_modified_time(project_id, dataset_id, table_id):
                credentials = service_account.Credentials.from_service_account_info(key)
                client = bigquery.Client(credentials=credentials, project=key['project_id'])
                table_ref = f"{project_id}.{dataset_id}.{table_id}"
                table = client.get_table(table_ref)

                last_modified_timestamp = table.modified
                last_modified_time = datetime.fromtimestamp(
                    last_modified_timestamp.timestamp(), tz=timezone.utc)

                return str(last_modified_time)

            def compile_results(df, columns_dict):
                missing_columns = columncheck(df, columns_dict)
                rows = getshape(df, 0)
                columns = getshape(df, 1)
                total_sales_amt = gettotal(df, 'pos_sales_amt')
                total_unit_sales_qty = gettotal(df, 'pos_unit_sales_qty')
                last_refresh = get_table_last_modified_time(
                    project_details["project_id"], project_details["dataset_id"], project_details["table_id"])

                missing_columns_str = ', '.join(
                    missing_columns) if missing_columns else ''

                results = {
                    'Country': [countryname],  # Convert scalar to list
                    'Missing Columns': [missing_columns_str],  # Keep as a list
                    'Number of Rows': [rows],  # Convert scalar to list
                    'Number of Columns': [float(columns)],  # Convert scalar to list
                    'Total Sales Amount': [float(total_sales_amt)], # Convert scalar to list
                    'Total Unit Sales Quantity': [float(total_unit_sales_qty)], # Keep as list of strings
                    'Check Last Refresh': [str(last_refresh)]
                }

                result_df = pd.DataFrame(results)
                return result_df

            # Section 2

            print("Loading Section 2 Functions ✅")

            # 1. Fucntion - Identify all Null Rows
            def getnullrows(data, checkname):
                subset_columns = section2a[checkname]
                result = []
                
                for index, row in data.iterrows():
                    null_cols = [col for col in subset_columns if pd.isnull(row[col])]
                    
                    if null_cols:
                        result.append({
                            'Record ID': index,
                            'Country': countryname,
                            'Sales Date': row['sales_date'],
                            'Error Type': 'Null Value Error',
                            'Columns': ', '.join(null_cols), 
                            'Sales Amount': row['pos_sales_amt'],
                            'Sales Quantity': row['pos_unit_sales_qty'],
                            'Store Key': row['store_key'],
                            'Manufacturer Name': row['mfgr_name'],
                            'Product Key': row['prod_key'],
                            'Retailer Name': row['retailer_name'],
                            'Retailer ID': row['retailer_id'],
                            'Sales Store Key': row['sales_store_key_reference'],
                            'Country Code': row['country_code'],
                            'Retailer Store ID': row['retailer_store_id'],
                        })

                result_df = pd.DataFrame(result)
                return result_df

            # Section 2
            # 1. Fucntion - Identify all Null Rows
            def getnanrows(data, checkname):
                subset_columns = section2a[checkname]
                result = []
                
                for index, row in data.iterrows():
                    nan_cols = [col for col in subset_columns if pd.isna(row[col])]
                    
                    if nan_cols:
                        result.append({
                            'Record ID': index,
                            'Country': countryname,
                            'Sales Date': row['sales_date'],
                            'Error Type': 'NaN Value Error',
                            'Columns': ', '.join(nan_cols),
                            'Sales Amount': row['pos_sales_amt'],
                            'Sales Quantity': row['pos_unit_sales_qty'],
                            'Store Key': row['store_key'],
                            'Manufacturer Name': row['mfgr_name'],
                            'Product Key': row['prod_key'],
                            'Retailer Name': row['retailer_name'],
                            'Retailer ID': row['retailer_id'],
                            'Sales Store Key': row['sales_store_key_reference'],
                            'Country Code': row['country_code'],
                            'Retailer Store ID': row['retailer_store_id'],
                        })

                result_df = pd.DataFrame(result)
                return result_df

            # 3. Fucntion - Identify all Duplicates
            def getduplicates(data, checkname):
                columns = section2a[checkname]
                result = []
                duplicates = data[data.duplicated(columns, keep=False)]

                for idx, row in duplicates.iterrows():
                    record_details = {
                        'Record ID': idx,
                        'Country': countryname,
                        'Sales Date': row['sales_date'],
                        'Error Type': 'Duplicate Value error',
                        'Columns': ', '.join(columns),
                        'Sales Amount': row['pos_sales_amt'],
                        'Sales Quantity': row['pos_unit_sales_qty'],
                        'Store Key': row['store_key'],
                        'Manufacturer Name': row['mfgr_name'],
                        'Product Key': row['prod_key'],
                        'Retailer Name': row['retailer_name'],
                        'Retailer ID': row['retailer_id'],
                        'Sales Store Key': row['sales_store_key_reference'],
                        'Country Code': row['country_code'],
                        'Retailer Store ID': row['retailer_store_id'],
                    }
                    result.append(record_details)

                result_df = pd.DataFrame(result)
                return result_df

            # 4. Fucntion - Checking for any Non-Display or Special characters
            def specialchars(s):
                non_display_characters = re.compile(r'[\x00-\x1F\x7F-\x9F]')
                if pd.isnull(s):
                    return 0
                if isinstance(s, str):
                    return len(non_display_characters.findall(s))
                else:
                    return 0

            def getspecialchars(data, checkname):
                subset_columns = section2a[checkname]
                subset_df = data[subset_columns]
                result = []
                non_display_counts = subset_df.applymap(specialchars)

                for idx, row in non_display_counts.iterrows():
                    if row.sum() > 0:
                        record_details = {
                            'Record ID': idx,
                            'Country': countryname,
                            'Sales Date': row['sales_date'],
                            'Error Type': 'Non-display Characters',
                            'Columns': ', '.join(row[row > 0].index),
                            'Sales Amount': row['pos_sales_amt'],
                            'Sales Quantity': row['pos_unit_sales_qty'],
                            'Store Key': row['store_key'],
                            'Manufacturer Name': row['mfgr_name'],
                            'Product Key': row['prod_key'],
                            'Retailer Name': row['retailer_name'],
                            'Retailer ID': row['retailer_id'],
                            'Sales Store Key': row['sales_store_key_reference'],
                            'Country Code': row['country_code'],
                            'Retailer Store ID': row['retailer_store_id']
                        }
                        result.append(record_details)

                result_df = pd.DataFrame(result)
                return result_df

            # 5. Check negatives (outliers) in Sales KPIs
            def checknegatives(data, checkname):
                columnnames = section2a[checkname]
                result = []

                for columnname in columnnames:
                    data[columnname] = data[columnname].replace([np.nan, float('inf'), float('-inf')], 0)
                    data[columnname] = pd.to_numeric(data[columnname], errors='coerce', downcast='integer')
                    
                    for idx, row in data.iterrows():
                        if row[columnname] < 0:
                            record_details = {
                                'Record ID': idx,
                                'Country': countryname,
                                'Sales Date': row['sales_date'],
                                'Error Type': 'Negative Values',
                                'Columns': columnname,
                                'Sales Amount': row['pos_sales_amt'],
                                'Sales Quantity': row['pos_unit_sales_qty'],
                                'Store Key': row['store_key'],
                                'Manufacturer Name': row['mfgr_name'],
                                'Product Key': row['prod_key'],
                                'Retailer Name': row['retailer_name'],
                                'Retailer ID': row['retailer_id'],
                                'Sales Store Key': row['sales_store_key_reference'],
                                'Country Code': row['country_code'],
                                'Retailer Store ID': row['retailer_store_id']
                            }
                            result.append(record_details)

                result_df = pd.DataFrame(result)
                return result_df

            # 6. Function for 0 sales = 0 quantity
            def pythonmatchsalesquantity(data, checkname):
                columnnames = section2a[checkname]
                column1 = columnnames[0]
                column2 = columnnames[1]

                data_copy = data.copy()

                # Convert the columns in the copied DataFrame to Decimal to maintain precision
                data_copy[column1] = data_copy[column1].apply(
                    lambda x: Decimal(x) if pd.notna(x) else None)
                data_copy[column2] = data_copy[column2].apply(
                    lambda x: Decimal(x) if pd.notna(x) else None)
                mismatched_rows = data_copy[((data_copy[column1] == Decimal(0)) & (data_copy[column2].notna()) & (data_copy[column2] != Decimal(0))) |
                                            ((data_copy[column1].notna()) & (data_copy[column1] != Decimal(0)) & (data_copy[column2] == Decimal(0)))]

                result = []

                for idx, row in mismatched_rows.iterrows():
                    record_details = {
                        'Record ID': idx,
                        'Country': countryname,
                        'Sales Date': row['sales_date'],
                        'Error Type': 'Sales & Quantity Mismatch error',
                        'Columns': [column1, column2],
                        'Sales Amount': row['pos_sales_amt'],
                        'Sales Quantity': row['pos_unit_sales_qty'],
                        'Store Key': row['store_key'],
                        'Manufacturer Name': row['mfgr_name'],
                        'Product Key': row['prod_key'],
                        'Retailer Name': row['retailer_name'],
                        'Retailer ID': row['retailer_id'],
                        'Sales Store Key': row['sales_store_key_reference'],
                        'Country Code': row['country_code'],
                        'Retailer Store ID': row['retailer_store_id']
                    }
                    result.append(record_details)

                result_df = pd.DataFrame(result)
                return result_df

            # 6. Function for 0 sales = 0 quantity

            def matchsalesquantity(data, checkname):
                real_df = data
                columnnames = section2a[checkname]
                column1 = columnnames[0]
                column2 = columnnames[1]

                mismtach_query = f"""
                WITH Union_Transactions AS (
                SELECT 
                    ROW_NUMBER() OVER() AS record_id,
                    *
                FROM (
                    SELECT *
                    FROM {view1}
                    UNION ALL
                    SELECT *
                    FROM {view2}
                    {date_limit}
                ) AS combined_transactions
                )
                SELECT 
                record_id,
                *
                FROM Union_Transactions
                WHERE ({column2} = 0 AND {column1} != 0) 
                OR ({column1} = 0 AND {column2} != 0);
                """

                query_job2 = client.query(mismtach_query)
                print("Query created for BigQuery ✅")
                mismtach_data = query_job2.to_dataframe()

                # Convert columns to float
                mismtach_data['pos_sales_amt'] = mismtach_data['pos_sales_amt'].astype(
                    float)
                mismtach_data['pos_unit_sales_qty'] = mismtach_data['pos_unit_sales_qty'].replace(
                    '0E-9', 0).astype(float)

                result = []

                for idx, row in mismtach_data.iterrows():
                    # Check if the row exists in the original dataframe
                    if real_df[
                        (real_df['sales_date'] == row['sales_date']) &
                        (real_df['pos_sales_amt'] == row['pos_sales_amt']) &
                        (real_df['pos_unit_sales_qty'] == row['pos_unit_sales_qty']) &
                        (real_df['store_key'] == row['store_key']) &
                        (real_df['mfgr_name'] == row['mfgr_name']) &
                        (real_df['prod_key'] == row['prod_key']) &
                        (real_df['retailer_name'] == row['retailer_name']) &
                        (real_df['sales_store_key_reference'] == row['sales_store_key_reference']) &
                        (real_df['country_code'] == row['country_code'])
                    ].empty == False:
                        record_details = {
                            'Record ID': row['record_id'],
                            'Country': countryname,
                            'Sales Date': row['sales_date'],
                            'Error Type': 'Sales & Quantity Mismatch error',
                            'Columns': [column1, column2],
                            'Sales Amount': row['pos_sales_amt'],
                            'Sales Quantity': row['pos_unit_sales_qty'],
                            'Store Key': row['store_key'],
                            'Manufacturer Name': row['mfgr_name'],
                            'Product Key': row['prod_key'],
                            'Retailer Name': row['retailer_name'],
                            'Retailer ID': row['retailer_id'],
                            'Sales Store Key': row['sales_store_key_reference'],
                            'Country Code': row['country_code'],
                            'Retailer Store ID': row['retailer_store_id']
                        }
                        result.append(record_details)

                result_df = pd.DataFrame(result)
                return result_df

            # 7. Check for Odd Data Types

            def datatype_check(data, column_name):
                data_types = data[column_name].apply(type)
                majority_type = data_types.mode()[0]
                output = data[data_types != majority_type].index.tolist()
                return output

            def getanomalousdatatypes(data, checkname):
                anomalous_indices = []
                subset_columns = section2a[checkname]
                subset_df = data[subset_columns]
                subset_df = pd.DataFrame(subset_df)

                for column_name in subset_df.columns:
                    indices = datatype_check(subset_df, column_name)
                    for idx, row in indices:
                        record_details = {
                            'Record ID': idx,
                            'Country': countryname,
                            'Sales Date': row['sales_date'],
                            'Error Type': 'Data Type Mismatch error',
                            'Columns': column_name,
                            'Sales Amount': row['pos_sales_amt'],
                            'Sales Quantity': row['pos_unit_sales_qty'],
                            'Store Key': row['store_key'],
                            'Manufacturer Name': row['mfgr_name'],
                            'Product Key': row['prod_key'],
                            'Retailer Name': row['retailer_name'],
                            'Retailer ID': row['retailer_id'],
                            'Sales Store Key': row['sales_store_key_reference'],
                            'Country Code': row['country_code'],
                            'Retailer Store ID': row['retailer_store_id']
                        }
                        anomalous_indices.append(record_details)

                result_df = pd.DataFrame(anomalous_indices)
                return result_df

            def checkint(data, checkname):
                result = []
                columns = section2a[checkname]
                for column in columns:
                    if not pd.api.types.is_numeric_dtype(data[column]):
                        record_details = {
                            'Record ID': "",
                            'Country': countryname,
                            'Record Value': "",
                            'Error Type': 'Column not Numeric',
                            'Columns': column
                        }
                        result.append(record_details)

                result_df = pd.DataFrame(result)
                return result_df

            # 9. Check for Country correctness

            def checkcountrycode(data, checkname):
                columnname = section2b[checkname]
                column = columnname[0]
                result = []

                for idx, row in data.iterrows():
                    country_code = row[column] if column in row else None
                    if country_code != code:
                        record_details = {
                            'Record ID': idx,
                            'Country': countryname,
                            'Sales Date': row['sales_date'],
                            'Error Type': 'Country Code Invalid error',
                            'Columns': column,
                            'Sales Amount': row['pos_sales_amt'],
                            'Sales Quantity': row['pos_unit_sales_qty'],
                            'Store Key': row['store_key'],
                            'Manufacturer Name': row['mfgr_name'],
                            'Product Key': row['prod_key'],
                            'Retailer Name': row['retailer_name'],
                            'Retailer ID': row['retailer_id'],
                            'Sales Store Key': row['sales_store_key_reference'],
                            'Country Code': row['country_code'],
                            'Retailer Store ID': row['retailer_store_id']
                        }
                        result.append(record_details)

                result_df = pd.DataFrame(result)
                return result_df

            # Section 2
            print("Loading Section 3 Functions ✅")

            # 10. Fucntion to check for invalid dates

            def checkdate(data, checkname):
                columnname = section2b[checkname]
                date_column = columnname[0]

                output = []
                wrong_dates = []
                current_date = datetime.now()

                for index, row in data.iterrows():
                    date = row[date_column]
                    try:
                        date = pd.to_datetime(date)
                        if date > current_date:
                            output.append({
                                'Record ID': index,
                                'Country': countryname,
                                'Sales Date': row['sales_date'],
                                'Error Type': 'Invalid Date error',
                                'Columns': date_column,
                                'Sales Amount': row['pos_sales_amt'],
                                'Sales Quantity': row['pos_unit_sales_qty'],
                                'Store Key': row['store_key'],
                                'Manufacturer Name': row['mfgr_name'],
                                'Product Key': row['prod_key'],
                                'Retailer Name': row['retailer_name'],
                                'Retailer ID': row['retailer_id'],
                                'Sales Store Key': row['sales_store_key_reference'],
                                'Country Code': row['country_code'],
                                'Retailer Store ID': row['retailer_store_id']
                            })
                    except Exception as e:
                        wrong_dates.append(date)

                result_df = pd.DataFrame(output)
                return result_df

            # Section 3

            # 1. Check for Retailer Name and ID Mapping
            def checkretailerkey(data, checkname):
                if section3[checkname] == "yes":

                    result = []

                    for idx, row in data.iterrows():
                        if retailers_dict.get(row['retailer_id']) != row['retailer_name']:
                            record_details = {
                                'Record ID': idx,
                                'Country': countryname,
                                'Sales Date': row['sales_date'],
                                'Error Type': 'Retailer Name Mismatch error',
                                'Columns': 'retailer_id, retailer_name',
                                'Sales Amount': row['pos_sales_amt'],
                                'Sales Quantity': row['pos_unit_sales_qty'],
                                'Store Key': row['store_key'],
                                'Manufacturer Name': row['mfgr_name'],
                                'Product Key': row['prod_key'],
                                'Retailer Name': row['retailer_name'],
                                'Retailer ID': row['retailer_id'],
                                'Sales Store Key': row['sales_store_key_reference'],
                                'Country Code': row['country_code'],
                                'Retailer Store ID': row['retailer_store_id']
                            }
                            result.append(record_details)

                    result_df = pd.DataFrame(result)
                    return result_df

            # 2. Check regex for sales_store_key_reference = retailer_id + country_code + retailer_store_id
            def checksalesstorekeyreference(data, checkname):
                finalcombination = False
                TH_expection = ["retail sales_Offline", "retail sales_Online","B2B sales_B2B-Don Jai", "B2B sales_B2B-Non Don Jai"]
                if section3[checkname] == "yes":
                    result = []
                    for index, row in data.iterrows():
                        if row['retailer_id'] == 2002:
                            for expcept_string in TH_expection:
                                combined_string = f"{row['country_code']}_{row['retailer_id']}_{row['retailer_store_id']}_{expcept_string}"
                                if combined_string != row['sales_store_key_reference']:
                                    combination = False
                                    if combination == True:
                                        finalcombination = True
                        if finalcombination != False:
                            record_details = {
                                'Record ID': index,
                                'Country': countryname,
                                'Sales Date': row['sales_date'],
                                'Error Type': 'Sales Store Key error',
                                'Columns': ['country_code', 'retailer_id', 'retailer_store_id'],
                                'Sales Amount': row['pos_sales_amt'],
                                'Sales Quantity': row['pos_unit_sales_qty'],
                                'Store Key': row['store_key'],
                                'Manufacturer Name': row['mfgr_name'],
                                'Product Key': row['prod_key'],
                                'Retailer Name': row['retailer_name'],
                                'Retailer ID': row['retailer_id'],
                                'Sales Store Key': row['sales_store_key_reference'],
                                'Country Code': row['country_code'],
                                'Retailer Store ID': row['retailer_store_id'],
                            }
                            result.append(record_details)
                        else:
                            combined_string = f"{row['country_code']}_{row['retailer_id']}_{row['retailer_store_id']}"
                            if combined_string != row['sales_store_key_reference']:
                                record_details = {
                                    'Record ID': index,
                                    'Country': countryname,
                                    'Sales Date': row['sales_date'],
                                    'Error Type': 'Sales Store Key error',
                                    'Columns': ['country_code', 'retailer_id', 'retailer_store_id'],
                                    'Sales Amount': row['pos_sales_amt'],
                                    'Sales Quantity': row['pos_unit_sales_qty'],
                                    'Store Key': row['store_key'],
                                    'Manufacturer Name': row['mfgr_name'],
                                    'Product Key': row['prod_key'],
                                    'Retailer Name': row['retailer_name'],
                                    'Retailer ID': row['retailer_id'],
                                    'Sales Store Key': row['sales_store_key_reference'],
                                    'Country Code': row['country_code'],
                                    'Retailer Store ID': row['retailer_store_id'],
                                }
                                result.append(record_details)

                    result_df = pd.DataFrame(result)
                    return result_df

            # 3. Check regex for sales_prod_key_reference = retailer_id + country_code + mfgr_name + retailer_prod_id (PH_6003_MAGNOLIA INC_144302)

            def checksalesprodkeyreference(data, checkname):
                # pattern = re.compile(r'PH_\d{4}_\d{3}')
                result = []

                if section3[checkname] == "yes":
                    for index, row in data.iterrows():
                        combined_string = f"{row['country_code']}_{row['retailer_id']}_{row['mfgr_name']}_{row['retailer_prod_id']}"
                        if combined_string != row['sales_prod_key_reference']:
                            record_details = {
                                'Record ID': index,
                                'Country': countryname,
                                'Sales Date': row['sales_date'],
                                'Error Type': 'Sales Product Key error',
                                'Columns': ['country_code', 'retailer_id', 'mfgr_name', 'retailer_prod_id'],
                                'Sales Amount': row['pos_sales_amt'],
                                'Sales Quantity': row['pos_unit_sales_qty'],
                                'Store Key': row['store_key'],
                                'Manufacturer Name': row['mfgr_name'],
                                'Product Key': row['prod_key'],
                                'Retailer Name': row['retailer_name'],
                                'Retailer ID': row['retailer_id'],
                                'Sales Store Key': row['sales_store_key_reference'],
                                'Country Code': row['country_code'],
                                'Retailer Store ID': row['retailer_store_id'],
                            }
                            result.append(record_details)

                    result_df = pd.DataFrame(result)
                    return result_df

            # 4. Check missin retailers

            def missing_retailers(data, checkname):
                if checkname in section3 and section3[checkname].lower() == "yes":
                    data['sales_date'] = pd.to_datetime(data['sales_date'])

                    periods = {
                        'Week': 'W',
                        'Half-Week': '2D',
                        'Month': 'M',
                        'Quarter': 'Q',
                        'Year': 'Y'
                    }

                    # Reading previous results to compare and merge
                    credentials = service_account.Credentials.from_service_account_info(devkey)
                    output_file = output_file_country
                    try:
                        if excelmode == True:
                            previous_results = read_excel_from_gcs(bucket_name, output_file_country, sheet_name=0)
                        else:
                            credentials = service_account.Credentials.from_service_account_info(key)
                            client = bigquery.Client(credentials=credentials, project=key['project_id'])
                            query = """ SELECT * FROM `dev-amea-analyt-epos-svc-ea.amea_dq_epos_all.missing_retailer` """
                            query_job = client.query(query)
                            results = query_job.result()
                            previous_results = results.to_dataframe()
                    except:
                        previous_results = pd.DataFrame()

                    result = []

                    for retailer in data['retailer_name'].unique():
                        retailer_data = data[data['retailer_name'] == retailer]
                        for period_name, period_code in periods.items():
                            # Resample the data to the specified period
                            resampled_data = retailer_data.resample(
                                period_code, on='sales_date').size()

                            # Find periods where sales are zero, indicating missing data
                            missing_periods = resampled_data[resampled_data == 0].index

                            for missing_period in missing_periods:
                                period_start = missing_period

                                # Calculate the end of the period
                                if period_code == 'M':
                                    period_end = period_start + pd.offsets.MonthEnd(1)
                                elif period_code == 'Q':
                                    period_end = period_start + \
                                        pd.offsets.QuarterEnd(1)
                                elif period_code == 'Y':
                                    period_end = period_start + pd.offsets.YearEnd(1)
                                elif period_code == 'W':
                                    period_end = period_start + pd.offsets.Week(1)
                                elif period_code == '2D':
                                    period_end = period_start + pd.Timedelta(days=3)

                                days_missing = (period_end - period_start).days + 1

                                if not previous_results.empty:
                                    match = previous_results[
                                        (previous_results['Retailer Name'] == retailer) &
                                        (previous_results['Period End Date']
                                        == period_start)
                                    ]

                                    if not match.empty:
                                        previous_results.loc[match.index,
                                                            'Period End Date'] = period_end
                                        continue

                                result.append({
                                    'Country': countryname,
                                    'Retailer Name': retailer,
                                    'Missing Period': period_name,
                                    'Period Start Date': period_start,
                                    'Period End Date': period_end,
                                    'Days Missing': int(days_missing),
                                    'Error Type': 'Missing Retailer Error',
                                    'Columns': 'retailer_name',
                                })

                    new_results = pd.DataFrame(result)
                    if not previous_results.empty:
                        new_results = pd.concat(
                            [previous_results, new_results]).drop_duplicates()

                    return new_results
                else:
                    return pd.DataFrame()


            # 5. Check Missing products
            def missing_products(data, checkname):
                if section3[checkname].lower() == "yes":
                    data['sales_date'] = pd.to_datetime(data['sales_date'])

                    periods = {
                        'Week': 'W',
                        'Half-Week': '2D',
                        'Month': 'M',
                        'Quarter': 'Q',
                        'Year': 'Y'
                    }

                    # Reading previous results to compare and merge
                    output_file = output_file_country
                    try:
                        if excelmode == True:
                            previous_results = read_excel_from_gcs(bucket_name, output_file_country, sheet_name=0)
                        else:
                            credentials = service_account.Credentials.from_service_account_info(key)
                            client = bigquery.Client(credentials=credentials, project=key['project_id'])
                            query = """ SELECT * FROM `dev-amea-analyt-epos-svc-ea.amea_dq_epos_all.missing_retailer` """
                            query_job = client.query(query)
                            results = query_job.result()
                            previous_results = results.to_dataframe()
                    except:
                        previous_results = pd.DataFrame()

                    result = []

                    for product in data['prod_key'].unique():
                        product_data = data[data['prod_key'] == product]
                        for period_name, period_code in periods.items():
                            # Resample the data to the specified period
                            resampled_data = product_data.resample(
                                period_code, on='sales_date').size()

                            # Find periods where sales are zero, indicating missing data
                            missing_periods = resampled_data[resampled_data == 0].index

                            for missing_period in missing_periods:
                                period_start = missing_period

                                # Calculate the end of the period
                                if period_code == 'M':
                                    period_end = period_start + pd.offsets.MonthEnd(1)
                                elif period_code == 'Q':
                                    period_end = period_start + \
                                        pd.offsets.QuarterEnd(1)
                                elif period_code == 'Y':
                                    period_end = period_start + pd.offsets.YearEnd(1)
                                elif period_code == 'W':
                                    period_end = period_start + pd.offsets.Week(1)
                                elif period_code == '2D':
                                    period_end = period_start + pd.Timedelta(days=3)

                                days_missing = (period_end - period_start).days + 1

                                if not previous_results.empty:
                                    match = previous_results[
                                        (previous_results['Product Name'] == product) &
                                        (previous_results['Period End Date']== period_start)
                                    ]

                                    if not match.empty:
                                        previous_results.loc[match.index,'Period End Date'] = period_end
                                        continue

                                result.append({
                                    'Country': countryname,
                                    'Product Name': product,
                                    'Missing Period': period_name,
                                    'Period Start Date': period_start,
                                    'Period End Date': period_end,
                                    'Days Missing': days_missing,
                                    'Error Type': 'Missing Product Error',
                                    'Columns': 'product_name',
                                })

                    new_results = pd.DataFrame(result)
                    if not previous_results.empty:
                        new_results = pd.concat(
                            [previous_results, new_results]).drop_duplicates()

                    return new_results
                else:
                    return pd.DataFrame()

            # 6. Check Sales Drop - Retailer

            def check_significant_sales_drop_store(data, checkname, country_value):
                if section3[checkname] == "yes":
                    data['sales_date'] = pd.to_datetime(data['sales_date'])
                    data['pos_sales_amt'] = data['pos_sales_amt'].astype(float)

                    data_with_period = data.copy()
                    data_with_period['year_month'] = data_with_period['sales_date'].dt.to_period(
                        'M')

                    avg_sales = data_with_period.groupby(['retailer_store_id', 'year_month'])[
                        'pos_sales_amt'].mean().reset_index()
                    avg_sales['prev_avg_sales'] = avg_sales.groupby(
                        'retailer_store_id')['pos_sales_amt'].shift(1)

                    data_with_avg = data_with_period.merge(avg_sales[['retailer_store_id', 'year_month', 'prev_avg_sales']],
                                                        on=['retailer_store_id', 'year_month'], how='left')
                    data_with_avg['prev_avg_sales'] = data_with_avg['prev_avg_sales'].astype(
                        float)

                    data_with_avg['drop_percent'] = (
                        data_with_avg['prev_avg_sales'] - data_with_avg['pos_sales_amt']) / data_with_avg['prev_avg_sales'] * 100
                    data_with_avg['significant_drop'] = data_with_avg['drop_percent'] > 90

                    # Check for months with no sales following a significant drop
                    significant_drops = data_with_avg[data_with_avg['significant_drop']].copy(
                    )
                    significant_drops['next_month'] = significant_drops['year_month'] + 1

                    no_sales_after_drop = significant_drops.merge(data_with_avg[['prod_key', 'year_month', 'pos_sales_amt']],
                                                                left_on=[
                                                                    'prod_key', 'next_month'],
                                                                right_on=[
                                                                    'prod_key', 'year_month'],
                                                                how='left',
                                                                suffixes=('', '_next'))

                    no_sales_after_drop = no_sales_after_drop[no_sales_after_drop['pos_sales_amt_next'] == 0]

                    # Sort by product key and sales date, then drop duplicates to keep the first occurrence
                    no_sales_after_drop = no_sales_after_drop.sort_values(
                        by=['retailer_store_id', 'sales_date']).drop_duplicates(subset=['retailer_store_id'])

                    # Add a new column 'Country' with the provided value
                    no_sales_after_drop['Country'] = country_value

                    # Rename the columns
                    no_sales_after_drop = no_sales_after_drop.rename(columns={
                        'retailer_store_id': 'Retailer Store ID',
                        'sales_date': 'Sales Date',
                        'year_month': 'Year-Month',
                        'pos_sales_amt': 'Sales Amount',
                        'prev_avg_sales': 'Previous Average Sale',
                        'drop_percent': 'Drop Percentage',
                        'Error Type': 'Dormant Store'
                    })

                return no_sales_after_drop[['Country', 'Retailer Store ID', 'Sales Date', 'Year-Month', 'Sales Amount', 'Previous Average Sale', 'Drop Percentage']]

            # 7. Check Sales Drop - Product

            def check_significant_sales_drop_product(data, checkname, country_value):
                if section3[checkname] == "yes":
                    data['sales_date'] = pd.to_datetime(data['sales_date'])
                    data['pos_sales_amt'] = data['pos_sales_amt'].astype(float)

                    data_with_period = data.copy()
                    data_with_period['year_month'] = data_with_period['sales_date'].dt.to_period(
                        'M')

                    avg_sales = data_with_period.groupby(['prod_key', 'year_month'])[
                        'pos_sales_amt'].mean().reset_index()
                    avg_sales['prev_avg_sales'] = avg_sales.groupby(
                        'prod_key')['pos_sales_amt'].shift(1)

                    data_with_avg = data_with_period.merge(avg_sales[['prod_key', 'year_month', 'prev_avg_sales']],
                                                        on=['prod_key', 'year_month'], how='left')
                    data_with_avg['prev_avg_sales'] = data_with_avg['prev_avg_sales'].astype(
                        float)

                    data_with_avg['drop_percent'] = (
                        data_with_avg['prev_avg_sales'] - data_with_avg['pos_sales_amt']) / data_with_avg['prev_avg_sales'] * 100
                    data_with_avg['significant_drop'] = data_with_avg['drop_percent'] > 90

                    # Check for months with no sales following a significant drop
                    significant_drops = data_with_avg[data_with_avg['significant_drop']].copy(
                    )
                    significant_drops['next_month'] = significant_drops['year_month'] + 1

                    no_sales_after_drop = significant_drops.merge(data_with_avg[['prod_key', 'year_month', 'pos_sales_amt']],
                                                                left_on=[
                                                                    'prod_key', 'next_month'],
                                                                right_on=[
                                                                    'prod_key', 'year_month'],
                                                                how='left',
                                                                suffixes=('', '_next'))

                    no_sales_after_drop = no_sales_after_drop[no_sales_after_drop['pos_sales_amt_next'] == 0]

                    # Sort by product key and sales date, then drop duplicates to keep the first occurrence
                    no_sales_after_drop = no_sales_after_drop.sort_values(
                        by=['prod_key', 'sales_date']).drop_duplicates(subset=['prod_key'])

                    # Add a new column 'Country' with the provided value
                    no_sales_after_drop['Country'] = country_value

                    # Rename the columns
                    no_sales_after_drop = no_sales_after_drop.rename(columns={
                        'prod_key': 'Product Key',
                        'sales_date': 'Sales Date',
                        'year_month': 'Year-Month',
                        'pos_sales_amt': 'Sales Amount',
                        'prev_avg_sales': 'Previous Average Sale',
                        'drop_percent': 'Drop Percentage',
                        'Error Type': 'Dormant Store'
                    })

                return no_sales_after_drop[['Country', 'Product Key', 'Sales Date', 'Year-Month', 'Sales Amount', 'Previous Average Sale', 'Drop Percentage']]

            # 8. Month on Month Growth

            def calculate_monthly_sales_growth(sales_column, data, checkname):
                if section3[checkname] == "yes":
                    data['sales_date'] = pd.to_datetime(data['sales_date'])
                    data['YearMonth'] = data['sales_date'].dt.to_period('M')
                    monthly_sales = data.groupby('YearMonth')[
                        sales_column].sum().reset_index()
                    monthly_sales['Growth Percentage'] = monthly_sales[sales_column].pct_change(
                    ) * 100
                    monthly_sales['Month'] = monthly_sales['YearMonth'].astype(str)
                    monthly_sales['Month_Name'] = monthly_sales['YearMonth'].dt.strftime(
                        '%B, %Y')
                    output = monthly_sales[['Month', sales_column,
                                            'Growth Percentage', 'Month_Name']]

                    output.columns = ['Month', 'Month Sales',
                                    'Growth Percentage', 'Month_Name']
                    output['Month Sales'] = output['Month Sales'].apply(
                        lambda x: f"{x:,.2f}")

                return output

            # 9. Month-Year Comparison
            def monthyeargrowth(data, sales_amt_column, sales_qty_column, countryname, view1, view2, checkname):
                if section3.get(checkname, '').lower() == "yes":
                    credentials = service_account.Credentials.from_service_account_info(key)
                    client = bigquery.Client(credentials=credentials, project=key['project_id'])

                    # Define the SQL query
                    monthgrowthquery = f"""
                    WITH unified_data AS (
                        SELECT 
                            DATE_TRUNC(sales_date, MONTH) AS Month,
                            {sales_amt_column} AS pos_sales_amt, 
                            {sales_qty_column} AS pos_unit_sales_qty,
                            '{countryname}' AS Country
                        FROM 
                            `{view1}`
                        WHERE 
                            sales_date BETWEEN '2022-01-01' AND CURRENT_DATE()
                        
                        UNION ALL
                        
                        SELECT 
                            DATE_TRUNC(sales_date, MONTH) AS Month,
                            {sales_amt_column} AS pos_sales_amt, 
                            {sales_qty_column} AS pos_unit_sales_qty,
                            '{countryname}' AS Country
                        FROM 
                            `{view2}`
                        WHERE 
                            sales_date BETWEEN '2022-01-01' AND CURRENT_DATE()
                    )

                    -- Now perform the aggregation on the unified data
                    SELECT 
                        Month,
                        SUM(pos_sales_amt) AS Total_Sales,
                        SUM(pos_unit_sales_qty) AS Total_Quantity,
                        '{countryname}' AS Country
                    FROM 
                        unified_data
                    GROUP BY 
                        Month, Country
                    ORDER BY 
                        Month;
                    """

                    credentials = service_account.Credentials.from_service_account_info(key)
                    client = bigquery.Client(credentials=credentials,project=key['project_id'])
                    query_job = client.query(monthgrowthquery)
                    results = query_job.result()
                    df = results.to_dataframe()

                    return df
            

            # 9. Sales Growth Comparison
            def monthyeargrowth(data, sales_amt_column, sales_qty_column, countryname, view1, view2, checkname):
                if section3.get(checkname, '').lower() == "yes":
                    credentials = service_account.Credentials.from_service_account_info(key)
                    client = bigquery.Client(credentials=credentials, project=key['project_id'])

                    # Define the SQL query
                    monthgrowthquery = f"""
                    WITH unified_data AS (
                        SELECT 
                            DATE_TRUNC(sales_date, MONTH) AS Month,
                            {sales_amt_column} AS pos_sales_amt, 
                            {sales_qty_column} AS pos_unit_sales_qty,
                            '{countryname}' AS Country
                        FROM 
                            `{view1}`
                        WHERE 
                            sales_date BETWEEN '2022-01-01' AND CURRENT_DATE()
                        
                        UNION ALL
                        
                        SELECT 
                            DATE_TRUNC(sales_date, MONTH) AS Month,
                            {sales_amt_column} AS pos_sales_amt, 
                            {sales_qty_column} AS pos_unit_sales_qty,
                            '{countryname}' AS Country
                        FROM 
                            `{view2}`
                        WHERE 
                            sales_date BETWEEN '2022-01-01' AND CURRENT_DATE()
                    )

                    -- Now perform the aggregation on the unified data
                    SELECT 
                        Month,
                        SUM(pos_sales_amt) AS Total_Sales,
                        SUM(pos_unit_sales_qty) AS Total_Quantity,
                        '{countryname}' AS Country
                    FROM 
                        unified_data
                    GROUP BY 
                        Month, Country
                    ORDER BY 
                        Month;
                    """

                    credentials = service_account.Credentials.from_service_account_info(key)
                    client = bigquery.Client(credentials=credentials,project=key['project_id'])
                    query_job = client.query(monthgrowthquery)
                    results = query_job.result()
                    df = results.to_dataframe()

                    return df


            # Calling - Section 1
            print("Processing Section 1 ✅")
            def section1checks(df, columns_dict):
                if section1['Check_All_Columns_Exist'] == 'yes':
                    missing_columns = columncheck(df, columns_dict)
                else:
                    missing_columns = ''

                if section1['Get_Total_Rows'] == 'yes':
                    rows = getshape(df, 0)
                else:
                    rows = ''

                if section1['Get_Total_Columns'] == 'yes':
                    columns = getshape(df, 1)
                else:
                    columns = ''

                if section1['Get_Total_Sales'] == 'yes':
                    total_sales_amt = gettotal(df, 'pos_sales_amt')
                else:
                    total_sales_amt = ''

                if section1['Get_Total_Quantity'] == 'yes':
                    total_unit_sales_qty = gettotal(df, 'pos_unit_sales_qty')
                else:
                    total_unit_sales_qty = ''

                if section1['Check_Last_Refresh'] == 'yes':
                    last_refresh = get_table_last_modified_time(
                        project_details["project_id"], project_details["dataset_id"], project_details["table_id"])
                else:
                    last_refresh = ''

                missing_columns_str = ', '.join(
                    missing_columns) if missing_columns else ''

                results = {
                    'Country': [str(countryname)],  # STRING field
                    'Missing Columns': [str(missing_columns_str)],  # STRING field
                    'Number of Rows': [int(rows) if rows else None],  # INTEGER field
                    'Number of Columns': [int(columns) if columns else None],  # INTEGER field
                    'Total Sales Amount': [float(total_sales_amt) if total_sales_amt else None],  # FLOAT field
                    'Total Unit Sales Quantity': [float(total_unit_sales_qty) if total_unit_sales_qty else None],  # FLOAT field
                    'Check Last Refresh': [str(last_refresh)]  # STRING field
                }

                result_df = pd.DataFrame(results)
                # print(result_df.dtypes)
                # print(result_df)
                return result_df

            # Section 2
            print("Processing Section 2 ✅")

            def section2checks(df):
                null_rows = getnullrows(df, 'Null_Check') 
                nan_rows = getnanrows(df, 'NaN_Check')
                duplicate_rows = getduplicates(df, 'Duplicate_Check')
                special_chars = getspecialchars(df, "Bad_Character_Check")
                negative_values = checknegatives(df, "Check_Negatives")
                sales_mismatch = matchsalesquantity(df, "Sales_Mismatch")
                odd_data_types = getanomalousdatatypes(df, 'Check_Odd_Data_Types')
                numeric_check = checkint(df, 'Numeric_Check')
                country_code_check = checkcountrycode(df, 'Country_Code_Check')
                date_check = checkdate(df, 'Invalid_Date_Check')
                checks_df = pd.concat([null_rows, nan_rows, duplicate_rows, special_chars, negative_values, sales_mismatch, odd_data_types, numeric_check, country_code_check, date_check]).reset_index(drop=True)

                # checks_df['Sales Date'] = pd.to_datetime(checks_df['Sales Date'], errors='coerce').dt.date
                checks_df['Sales Amount'] = pd.to_numeric(checks_df['Sales Amount'], errors='coerce')
                checks_df['Sales Quantity'] = pd.to_numeric(checks_df['Sales Quantity'], errors='coerce')
                checks_df['Store Key'] = checks_df['Store Key'].astype(str)
                checks_df['Product Key'] = checks_df['Product Key'].astype(str)
                checks_df['Retailer ID'] = pd.to_numeric(checks_df['Retailer ID'], errors='coerce')

                columns_to_cast_to_string = ['Country', 'Error Type', 'Columns', 'Manufacturer Name', 'Retailer Name', 'Sales Store Key', 'Country Code', 'Retailer Store ID', 'Sales Date']
                checks_df[columns_to_cast_to_string] = checks_df[columns_to_cast_to_string].astype(str)

                return checks_df

            # Section 3
            print("Processing Section 3 ✅")

            def section3checks(df):
                product_key_check = checkretailerkey(df, "Check_Product_Key")
                store_key_check = checksalesstorekeyreference(df, "Check_Store_Key")
                retailer_key_check = checksalesprodkeyreference(df, "Check_Retailer_Key")

                key_checks_df = pd.concat([product_key_check, store_key_check, retailer_key_check]).reset_index(drop=True)
                if not key_checks_df.empty:
                    key_checks_df['Sales Amount'] = pd.to_numeric(key_checks_df['Sales Amount'], errors='coerce')
                    key_checks_df['Sales Quantity'] = pd.to_numeric(key_checks_df['Sales Quantity'], errors='coerce')
                    key_checks_df['Store Key'] = key_checks_df['Store Key'].astype(str)
                    key_checks_df['Product Key'] = key_checks_df['Product Key'].astype(str)
                    key_checks_df['Retailer ID'] = pd.to_numeric(key_checks_df['Retailer ID'], errors='coerce')

                    columns_to_cast_to_string = ['Country', 'Error Type', 'Columns', 'Manufacturer Name', 'Retailer Name', 'Sales Store Key', 'Country Code', 'Retailer Store ID', 'Sales Date']
                    key_checks_df[columns_to_cast_to_string] = key_checks_df[columns_to_cast_to_string].astype(str)

                return key_checks_df

            def missingretailer(df):
                missing_retailer_check = missing_retailers(df, "Check_Missing_Retailer")
                columns_to_cast_to_string = ['Period End Date', 'Period Start Date']
                missing_retailer_check[columns_to_cast_to_string] = missing_retailer_check[columns_to_cast_to_string].astype(str)
                return missing_retailer_check

            def missingproduct(df):
                missing_product_check = missing_products(df, "Check_Missing_Product")
                missing_product_check['Days Missing'] = pd.to_numeric(missing_product_check['Days Missing'], errors='coerce')

                columns_to_cast_to_string = ['Country', 'Product Name', 'Missing Period', 'Period Start Date', 'Period End Date', 'Error Type', 'Columns']
                missing_product_check[columns_to_cast_to_string] = missing_product_check[columns_to_cast_to_string].astype(str)


                return missing_product_check

            def checkmonthyeargrowth(df):
                credentials = service_account.Credentials.from_service_account_info(key)
                client = bigquery.Client(credentials=credentials,project=key['project_id'])
                monthyear = monthyeargrowth(df, 'pos_sales_amt', 'pos_unit_sales_qty', countryname, view1, view2, "Check_Month_on_Month_Growth")
                monthyear['Total_Sales'] = pd.to_numeric(monthyear['Total_Sales'], errors='coerce')
                monthyear['Total_Quantity'] = pd.to_numeric(monthyear['Total_Quantity'], errors='coerce')
                columns_to_cast_to_string = ['Country', 'Month']
                monthyear[columns_to_cast_to_string] = monthyear[columns_to_cast_to_string].astype(str)

                return monthyear

            def checkdormantstore(df):
                dormant_store_check = check_significant_sales_drop_store(df, "Check_Dormant_Store", countryname)
                columns_to_cast_to_string = ['Country', 'Retailer Store ID', 'Year-Month', 'Sales Date']
                dormant_store_check[columns_to_cast_to_string] = dormant_store_check[columns_to_cast_to_string].astype(str)
                return dormant_store_check

            def checkdormantproduct(df):
                dormant_product_check = check_significant_sales_drop_product(df, "Check_Dormant_Product", countryname)
                columns_to_cast_to_string = ['Country', 'Product Key', 'Year-Month', 'Sales Date']
                dormant_product_check[columns_to_cast_to_string] = dormant_product_check[columns_to_cast_to_string].astype(str)
                return dormant_product_check

            print("Executing Checks Results ✅")
            section_1_results = section1checks(df, columnsdict)
            print("Processed Section 1 Results ✅")
            section_2_results = section2checks(df)
            print("Processed Section 2 Results ✅")
            section_3_results = section3checks(df)
            print("Processed Section 3 Results ✅")
            missingretailer_results = missingretailer(df)
            print("Processed Missing Retailer Results ✅")
            missingproduct_results = missingproduct(df)
            print("Processed Missing Product Results ✅")
            checkmonthonmonthresults = checkmonthyeargrowth(df)
            print("Processed Month on Month Results ✅")
            checkdormantstore_results = checkdormantstore(df)
            print("Processed Dormat Store Results ✅")
            checkdormantproduct_results = checkdormantproduct(df)
            print("Processed Dormat Product Results ✅")

            output_file = output_file_country
            if os.path.exists(output_file):
                os.remove(output_file)
                print(f"Existing file {output_file} has been deleted.")

            print("Compiling Final Output ✅")

            def drop_duplicates_safe(df):
                # Identify columns with unhashable types (e.g., lists)
                hashable_columns = [col for col in df.columns if df[col].apply(
                    lambda x: isinstance(x, (list, dict))).sum() == 0]
                # Drop duplicates considering only hashable columns
                return df.drop_duplicates(subset=hashable_columns)

            def remove_sheet_if_exists(writer, sheet_name):
                # Remove the sheet if it exists
                if sheet_name in writer.book.sheetnames:
                    std = writer.book[sheet_name]
                    writer.book.remove(std)

            def push_to_bigquery(dataframe, table_id):
                credentials = service_account.Credentials.from_service_account_info(devkey)
                client = bigquery.Client(credentials=credentials,project=devkey['project_id'])
                # Ensure all columns are converted to types compatible with BigQuery
                dataframe = dataframe.applymap(lambda x: str(x) if isinstance(x, (list, dict)) else x)

                try:
                    # Check if the table already exists
                    query = f"SELECT * FROM `{table_id}` WHERE country_code = '{code}' "
                    existing_data = client.query(query).to_dataframe()
                    print(table_id)
                    print(f"Data fetched from {table_id} for duplicate check.")
                except Exception as e:
                    # If the table does not exist, proceed without checking duplicates
                    print(f"Table {table_id} does not exist yet. Creating new table.")
                    existing_data = pd.DataFrame()

                # If the table exists and has data, check for any duplicates in the new data
                if not existing_data.empty:
                    # Check for any duplicate rows in the new data compared to the existing data      
                    duplicates = pd.merge(existing_data, dataframe, how='inner')

                    if not duplicates.empty:
                        print(f"Duplicates found! Data not pushed to {table_id}.")
                    else:
                        # No duplicates found, push the entire dataframe
                        job_config = bigquery.LoadJobConfig(
                            write_disposition=bigquery.WriteDisposition.WRITE_APPEND
                        )
                        job = client.load_table_from_dataframe(
                            dataframe, table_id, job_config=job_config)
                        job.result() 
                        print(f"Data successfully written to {table_id}\n")
                else:
                    # If table does not exist or is empty, push all the data
                    job_config = bigquery.LoadJobConfig(
                        write_disposition=bigquery.WriteDisposition.WRITE_APPEND
                    )
                    job = client.load_table_from_dataframe(
                        dataframe, table_id, job_config=job_config)
                    job.result()  # Wait for the job to complete
                    print(
                        f"Data successfully written to {table_id} (No existing data).")

            if excelmode == True:
                if os.path.exists(output_file):
                    book = load_workbook(output_file)
                    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                        writer.book = book

                        # Section 1 Results
                        if 'Section 1 Results' in writer.book.sheetnames:
                            existing_data = pd.read_excel(
                                output_file, sheet_name='Section 1 Results', engine='openpyxl')
                            section_1_results = pd.concat(
                                [existing_data, section_1_results])
                            section_1_results = drop_duplicates_safe(section_1_results)
                            remove_sheet_if_exists(writer, 'Section 1 Results')
                        section_1_results.to_excel(
                            writer, sheet_name='Section 1 Results', index=False)

                        # Section 2 Results
                        if 'Section 2 Results' in writer.book.sheetnames:
                            existing_data = pd.read_excel(
                                output_file, sheet_name='Section 2 Results', engine='openpyxl')
                            section_2_results = pd.concat(
                                [existing_data, section_2_results])
                            section_2_results = drop_duplicates_safe(section_2_results)
                            remove_sheet_if_exists(writer, 'Section 2 Results')
                        section_2_results.to_excel(
                            writer, sheet_name='Section 2 Results', index=False)

                        # Section 3 Results
                        if 'Section 3 Results' in writer.book.sheetnames:
                            existing_data = pd.read_excel(
                                output_file, sheet_name='Section 3 Results', engine='openpyxl')
                            section_3_results = pd.concat(
                                [existing_data, section_3_results])
                            section_3_results = drop_duplicates_safe(section_3_results)
                            remove_sheet_if_exists(writer, 'Section 3 Results')
                        section_3_results.to_excel(
                            writer, sheet_name='Section 3 Results', index=False)

                        # Missing Retailer Results
                        if 'Missing Retailer Results' in writer.book.sheetnames:
                            existing_data = pd.read_excel(
                                output_file, sheet_name='Missing Retailer Results', engine='openpyxl')
                            missingretailer_results = pd.concat(
                                [existing_data, missingretailer_results])
                            missingretailer_results = drop_duplicates_safe(
                                missingretailer_results)
                            remove_sheet_if_exists(writer, 'Missing Retailer Results')
                        missingretailer_results.to_excel(
                            writer, sheet_name='Missing Retailer Results', index=False)

                        # Missing Product Results
                        if 'Missing Product Results' in writer.book.sheetnames:
                            existing_data = pd.read_excel(
                                output_file, sheet_name='Missing Product Results', engine='openpyxl')
                            missingproduct_results = pd.concat(
                                [existing_data, missingproduct_results])
                            missingproduct_results = drop_duplicates_safe(
                                missingproduct_results)
                            remove_sheet_if_exists(writer, 'Missing Product Results')
                        missingproduct_results.to_excel(
                            writer, sheet_name='Missing Product Results', index=False)

                        # Month on Month Growth Results
                        if 'Month on Month Growth Results' in writer.book.sheetnames:
                            existing_data = pd.read_excel(
                                output_file, sheet_name='Month on Month Growth Results', engine='openpyxl')
                            checkmonthonmonthresults = pd.concat(
                                [existing_data, checkmonthonmonthresults])
                            checkmonthonmonthresults = drop_duplicates_safe(
                                checkmonthonmonthresults)
                            remove_sheet_if_exists(
                                writer, 'Month on Month Growth Results')
                        checkmonthonmonthresults.to_excel(
                            writer, sheet_name='Month on Month Growth Results', index=False)

                        # Dormant Store Results
                        if 'Dormant Store Results' in writer.book.sheetnames:
                            existing_data = pd.read_excel(
                                output_file, sheet_name='Dormant Store Results', engine='openpyxl')
                            checkdormantstore_results = pd.concat(
                                [existing_data, checkdormantstore_results])
                            checkdormantstore_results = drop_duplicates_safe(
                                checkdormantstore_results)
                            remove_sheet_if_exists(writer, 'Dormant Store Results')
                        checkdormantstore_results.to_excel(
                            writer, sheet_name='Dormant Store Results', index=False)

                        # Dormant Product Results
                        if 'Dormant Product Results' in writer.book.sheetnames:
                            existing_data = pd.read_excel(
                                output_file, sheet_name='Dormant Product Results', engine='openpyxl')
                            checkdormantproduct_results = pd.concat(
                                [existing_data, checkdormantproduct_results])
                            checkdormantproduct_results = drop_duplicates_safe(
                                checkdormantproduct_results)
                            remove_sheet_if_exists(writer, 'Dormant Product Results')
                        checkdormantproduct_results.to_excel(
                            writer, sheet_name='Dormant Product Results', index=False)

                else:
                    dataframe_dict = {
                        'Section 1 Results': section_1_results,
                        'Section 2 Results': section_2_results,
                        'Section 3 Results': section_3_results,
                        'Missing Retailer Results': missingretailer_results,
                        'Missing Product Results': missingproduct_results,
                        'Month on Month Growth Results': checkmonthonmonthresults,
                        'Dormant Store Results': checkdormantstore_results,
                        'Dormant Product Results': checkdormantproduct_results,
                    }

                # Write the dataframes to the output GCS path
                write_excel_to_gcs(dataframe_dict, bucket_name, output_file_country)

                print(f"Results have been saved to {output_file_country} on GCS")
                print("Successfully run all checks ✅")

            else:
                project_id = pushproject
                dataset_id = pushdataset

                credentials = service_account.Credentials.from_service_account_info(devkey)
                client = bigquery.Client(credentials=credentials,project=devkey['project_id'])

                section_1_table_id = f"{project_id}.{dataset_id}.section1_logs"
                if not section_1_results.empty:
                    print(f"Pushing section_1_results to table: {section_1_table_id}")
                    push_to_bigquery(section_1_results, section_1_table_id)

                section_2_table_id = f"{project_id}.{dataset_id}.section2_logs"
                if not section_2_results.empty:
                    print(f"Pushing section_2_results to table: {section_2_table_id}")
                    push_to_bigquery(section_2_results, section_2_table_id)
                
                section_3_table_id = f"{project_id}.{dataset_id}.section_2_logs"
                if not section_3_results.empty:
                    print(f"Pushing section_3_results to table: {section_3_table_id}")
                    push_to_bigquery(section_3_results, section_3_table_id)
                
                missing_retailer_table_id = f"{project_id}.{dataset_id}.missing_retailer"
                if not missingretailer_results.empty:
                    print(f"Pushing missingretailer_results to table: {missing_retailer_table_id}")
                    push_to_bigquery(missingretailer_results, missing_retailer_table_id)

                missing_product_table_id = f"{project_id}.{dataset_id}.missing_product"
                if not missingproduct_results.empty:
                    print(f"Pushing missingproduct_results to table: {missing_product_table_id}")
                    push_to_bigquery(missingproduct_results, missing_product_table_id)

                dormant_store_table_id = f"{project_id}.{dataset_id}.dormat_store"
                if not checkdormantstore_results.empty:
                    print(f"Pushing checkdormantstore_results to table: {dormant_store_table_id}")
                    push_to_bigquery(checkdormantstore_results, dormant_store_table_id)

                dormant_product_table_id = f"{project_id}.{dataset_id}.dormant_product"
                if not checkdormantproduct_results.empty:
                    print(f"Pushing checkdormantproduct_results to table: {dormant_product_table_id}")
                    push_to_bigquery(checkdormantproduct_results, dormant_product_table_id)
                
                mom_growth_table_id = f"{project_id}.{dataset_id}.monthyear_growth"
                if not checkmonthonmonthresults.empty:
                    print(f"Pushing checkmonthonmonthresults to table: {mom_growth_table_id}")
                    push_to_bigquery(checkmonthonmonthresults, mom_growth_table_id)

                print("Successfully pushed all data to BigQuery tables ✅")

            def updatedate_gcs(data, bucket_name, output_file, country_code):
                data['sales_date'] = pd.to_datetime(data['sales_date'])
                latest_sales_date = data['sales_date'].max()

                # Try to read the existing date.json from GCS, else create a new one
                try:
                    latest_date_data = read_json_from_gcs(bucket_name, output_file)
                except Exception:
                    latest_date_data = {"Latest Date": {}}

                latest_date_data["Latest Date"][country_code] = latest_sales_date.strftime("%Y-%m-%d")

                # Write the updated date.json back to GCS
                write_json_to_gcs(latest_date_data, bucket_name, output_file)
                return latest_date_data

            updatedate_gcs(df, bucket_name, 'dashboard/date.json', code)

        else:
            print("Wrong Country Passed")



    # Function to read JSON from GCS
    def read_json_from_gcs(bucket_name, file_path):
        credentials = service_account.Credentials.from_service_account_info(devkey)
        client = storage.Client(credentials=credentials, project=devkey['project_id'])
        bucket = client.bucket(bucket_name)
        blob = bucket.blob(file_path)
        json_data = json.loads(blob.download_as_string())
        print("")

        return json_data


    # Replace with GCS bucket and file path
    bucket_name = "dq_dashboard"
    date_file = "dashboard/date.json"

    credentials = service_account.Credentials.from_service_account_info(devkey)
    client = storage.Client(credentials=credentials,project=devkey['project_id'])
    date_variables = read_json_from_gcs(bucket_name, 'dashboard/date.json')

    countries = {
        "TH": {
            "Retailers_Dict": {
                2000: "LOTUS",
                2001: "TOPS",
                2002: "Big C",
                2003: "MAKRO",
                2004: "7ELEVEN"
            },
            "Country Code": "TH",
            "Country Name": "Thailand",
            "Input File": "dashboard/Input_Folder/TH_Input.xlsx",
            "Output File": "dashboard/TH_Output.xlsx",
            "Fact Tables": {
                "view1": "prd-amea-analyt-epos-svc-47.amea_p_epos_th.v_fct_epos_transactions_mth",
                "view2": "prd-amea-analyt-epos-svc-47.amea_p_epos_th.v_fct_epos_transactions_wk"
            },
            "project_id": "dev-amea-analyt-epos-svc-ea",
            "dataset_id": "amea_dq_epos_all",
            "columnsdict": {
                'retailer_id': [],
                'retailer_name': [],
                'prod_key': [],
                'store_key': [],
                'prod_hist_key': [],
                'store_hist_key': [],
                'country_code': [],
                'material_id': [],
                'material_name': [],
                'internal_item_id': [],
                'internal_item_name': [],
                'ean_code': [],
                'ean_name': [],
                'retailer_prod_id': [],
                'retailer_store_id': [],
                'retailer_store_name': [],
                'store_id': [],
                'sales_rep_id': [],
                'sales_date': [],
                'data_grain': [],
                'file_type': [],
                'segment_id': [],
                'segment_name': [],
                'is_innovation': [],
                'is_promo': [],
                'promo_year': [],
                'promo_week': [],
                'promo_description': [],
                'discount': [],
                'promo_definition': [],
                'pos_unit_sales_qty': [],
                'pos_sales_amt': [],
                'pos_sales_online_amt': [],
                'pos_volume_sales_qty': [],
                'shelf_price_amt': [],
                'price_start_date': [],
                'price_end_date': [],
                'delta_flag': [],
                'mfgr_name': [],
                'channel_name': [],
                'sales_store_key_reference': [],
                'sales_prod_key_reference': [],
                'source_system': [],
                'bq_insert_timestamp': [],
                'bq_update_timestamp': [],
                'bq_job_name': [],
            }
        },
        "ID": {
            "Retailers_Dict": {
                3000: "ALFAMIDI",
                3001: "INDOMARET",
                3002: "ALFAMART",
                3003: "Hypermart",
                3004: "LSI",
                3005: "YOGYA",
                3006: "YOMART"
            },
            "Country Code": "ID",
            "Country Name": "Indonesia",
            "Input File": "dashboard/Input_Folder/ID_Input.xlsx",
            "Output File": "dashboard/ID_Output.xlsx",
            "Fact Tables": {
                "view1": "prd-amea-analyt-epos-svc-47.amea_p_epos_id.v_fct_epos_transactions_mth",
                "view2": "prd-amea-analyt-epos-svc-47.amea_p_epos_id.v_fct_epos_transactions_mth"
            },
            "project_id": "dev-amea-analyt-epos-svc-ea",
            "dataset_id": "amea_dq_epos_all",
            "columnsdict": {
                'retailer_id': [],
                'retailer_name': [],
                'prod_key': [],
                'store_key': [],
                'prod_hist_key': [],
                'store_hist_key': [],
                'country_code': [],
                'material_id': [],
                'material_name': [],
                'internal_item_id': [],
                'internal_item_name': [],
                'ean_code': [],
                'ean_name': [],
                'retailer_prod_id': [],
                'retailer_store_id': [],
                'retailer_store_name': [],
                'store_id': [],
                'sales_rep_id': [],
                'sales_date': [],
                'data_grain': [],
                'file_type': [],
                'segment_id': [],
                'segment_name': [],
                'is_innovation': [],
                'is_promo': [],
                'promo_year': [],
                'promo_week': [],
                'promo_description': [],
                'discount': [],
                'promo_definition': [],
                'pos_unit_sales_qty': [],
                'pos_sales_amt': [],
                'pos_sales_online_amt': [],
                'pos_volume_sales_qty': [],
                'shelf_price_amt': [],
                'price_start_date': [],
                'price_end_date': [],
                'delta_flag': [],
                'mfgr_name': [],
                'channel_name': [],
                'sales_store_key_reference': [],
                'sales_prod_key_reference': [],
                'source_system': [],
                'bq_insert_timestamp': [],
                'bq_update_timestamp': [],
                'bq_job_name': [],
                'bq_updated_by': [],
                'bq_inserted_by': [],
                'transaction_key': []
            }
        },
        "MY": {
            "Retailers_Dict": {
                1000: "AEON",
                1001: "TRENDCELL",
                1002: "99SM"
            },
            "Country Code": "MY",
            "Country Name": "Malaysia",
            "Input File": "dashboard/Input_Folder/MY_Input.xlsx",
            "Output File": "dashboard/MY_Output.xlsx",
            "Fact Tables": {
                "view1": "prd-amea-analyt-epos-svc-47.amea_p_epos_my.v_fct_epos_transactions_day",
                "view2": "prd-amea-analyt-epos-svc-47.amea_p_epos_my.v_fct_epos_transactions_mth"
            },
            "project_id": "dev-amea-analyt-epos-svc-ea",
            "dataset_id": "amea_dq_epos_all",
            "columnsdict": {
                'retailer_id': [],
                'retailer_name': [],
                'prod_key': [],
                'store_key': [],
                'prod_hist_key': [],
                'store_hist_key': [],
                'country_code': [],
                'material_id': [],
                'material_name': [],
                'internal_item_id': [],
                'internal_item_name': [],
                'ean_code': [],
                'ean_name': [],
                'retailer_prod_id': [],
                'retailer_store_id': [],
                'retailer_store_name': [],
                'store_id': [],
                'sales_rep_id': [],
                'sales_date': [],
                'data_grain': [],
                'file_type': [],
                'segment_id': [],
                'segment_name': [],
                'is_innovation': [],
                'is_promo': [],
                'promo_year': [],
                'promo_week': [],
                'promo_description': [],
                'discount': [],
                'promo_definition': [],
                'pos_unit_sales_qty': [],
                'pos_sales_amt': [],
                'pos_sales_online_amt': [],
                'pos_volume_sales_qty': [],
                'shelf_price_amt': [],
                'price_start_date': [],
                'price_end_date': [],
                'delta_flag': [],
                'mfgr_name': [],
                'channel_name': [],
                'sales_store_key_reference': [],
                'sales_prod_key_reference': [],
                'source_system': [],
                'bq_insert_timestamp': [],
                'bq_update_timestamp': [],
                'bq_job_name': [],
                'bq_updated_by': [],
                'bq_inserted_by': [],
                'transaction_key': []
            }
        },
        "SG": {
            "Retailers_Dict": {
                5000: "SHENG SIONG",
                5001: "FAIRPRICE",
                5002: "7ELEVEN",
                5003: "COLD STORAGE",
                5004: "GIANT"
            },
            "Country Code": "SG",
            "Country Name": "Singapore",
            "Input File": "dashboard/Input_Folder/SG_Input.xlsx",
            "Output File": "dashboard/SG_Output.xlsx",
            "Fact Tables": {
                "view1": "prd-amea-analyt-epos-svc-47.amea_p_epos_sg.v_fct_epos_transactions_day",
                "view2": "prd-amea-analyt-epos-svc-47.amea_p_epos_sg.v_fct_epos_transactions_mth"
            },
            "project_id": "dev-amea-analyt-epos-svc-ea",
            "dataset_id": "amea_dq_epos_all",
            "columnsdict": {
                'retailer_id': [],
                'retailer_name': [],
                'prod_key': [],
                'store_key': [],
                'prod_hist_key': [],
                'store_hist_key': [],
                'country_code': [],
                'file_type': [],
                'delta_flag': [],
                'TRANSACTION_KEY': [],
                'material_id': [],
                'material_name': [],
                'internal_item_id': [],
                'internal_item_name': [],
                'ean_code': [],
                'ean_name': [],
                'retailer_prod_id': [],
                'retailer_store_id': [],
                'retailer_store_name': [],
                'store_id': [],
                'sales_rep_id': [],
                'sales_date': [],
                'data_grain': [],
                'segment_id': [],
                'segment_name': [],
                'is_innovation': [],
                'is_promo': [],
                'promo_year': [],
                'promo_week': [],
                'promo_description': [],
                'discount': [],
                'promo_definition': [],
                'pos_unit_sales_qty': [],
                'pos_sales_amt': [],
                'pos_sales_online_amt': [],
                'pos_volume_sales_qty': [],
                'shelf_price_amt': [],
                'price_start_date': [],
                'price_end_date': [],
                'mfgr_name': [],
                'channel_name': [],
                'sales_store_key_reference': [],
                'sales_prod_key_reference': [],
                'source_system': [],
                'bq_insert_timestamp': [],
                'bq_update_timestamp': [],
                'bq_job_name': [],
                'bq_updated_by': [],
                'bq_inserted_by': []
            }
        },
        "PH": {
            "Retailers_Dict": {
                6000: "Metro Gaisano",
                6009: "ROBINSON DEPARTMENT STORE",
                6010: "LCC",
                6020: "NCCC",
                6021: "PRINCE",
                6011: "TMP_SW_RE",
                6006: "SOUTHSTAR",
                6001: "GAISANO GRAND",
                6002: "CITIMART",
                6004: "UNCLE_JOHNS",
                6003: "WALTERMART",
                6005: "SUYSING",
                6008: "7ELEVEN",
                6007: "ROBINSONS_SMKT",
            },
            "Country Code": "PH",
            "Country Name": "Philippines",
            "Input File": "dashboard/Input_Folder/PH_Input.xlsx",
            "Output File": "dashboard/PH_Output.xlsx",
            "Fact Tables": {
                "view1": "prd-amea-analyt-epos-svc-47.amea_p_epos_ph.v_fct_epos_transactions_day_ph_dqdb",
                "view2": "prd-amea-analyt-epos-svc-47.amea_p_epos_ph.v_fct_epos_transactions_mth_ph_dqdb"
            },
            "project_id": "dev-amea-analyt-epos-svc-ea",
            "dataset_id": "amea_dq_epos_all",
            "columnsdict": {
                'retailer_id': [],
                'retailer_name': [],
                'prod_key': [],
                'store_key': [],
                'prod_hist_key': [],
                'store_hist_key': [],
                'country_code': [],
                'file_type': [],
                'delta_flag': [],
                'TRANSACTION_KEY': [],
                'material_id': [],
                'material_name': [],
                'internal_item_id': [],
                'internal_item_name': [],
                'ean_code': [],
                'ean_name': [],
                'retailer_prod_id': [],
                'retailer_store_id': [],
                'retailer_store_name': [],
                'store_id': [],
                'sales_rep_id': [],
                'sales_date': [],
                'data_grain': [],
                'segment_id': [],
                'segment_name': [],
                'is_innovation': [],
                'is_promo': [],
                'promo_year': [],
                'promo_week': [],
                'promo_description': [],
                'discount': [],
                'promo_definition': [],
                'pos_unit_sales_qty': [],
                'pos_sales_amt': [],
                'pos_sales_online_amt': [],
                'pos_volume_sales_qty': [],
                'shelf_price_amt': [],
                'price_start_date': [],
                'price_end_date': [],
                'mfgr_name': [],
                'channel_name': [],
                'sales_store_key_reference': [],
                'sales_prod_key_reference': [],
                'source_system': [],
                'bq_insert_timestamp': [],
                'bq_update_timestamp': [],
                'bq_job_name': [],
                'bq_updated_by': [],
                'bq_inserted_by': []
            }


        }
    }

    credentials = service_account.Credentials.from_service_account_info(devkey)
    client = bigquery.Client(credentials=credentials,project=devkey['project_id'])
    project_id = "dev-amea-analyt-epos-svc-ea"
    dataset_id = "amea_dq_epos_all"

    for country_code, details in countries.items():
        retailers_dict = details["Retailers_Dict"]
        output_file_country = details["Output File"]
        view1 = details["Fact Tables"]["view1"]
        view2 = details["Fact Tables"]["view2"]
        countryname = details["Country Name"]
        pushproject = details["project_id"]
        pushdataset = details["dataset_id"]
        input_file_country = details["Input File"]
        columnsdict = details["columnsdict"]
        date1 = date_variables["Latest Date"].get(country_code, None)

        if date1 is None:
            print(f"No date found for country {country_code} in JSON file. Skipping...")
            continue

        highest_date_query = f"""
            SELECT MAX(max_date) FROM (
                SELECT MAX(sales_date) AS max_date FROM {view1}
                UNION ALL
                SELECT MAX(sales_date) FROM {view2}
            ) AS combined_dates;"""

        main(bucket_name, retailers_dict, country_code, countryname, view1, view2, input_file_country, output_file_country, columnsdict, date1, pushproject, pushdataset)
        print(f"Processed data for country {country_code} ✨")
    print(f"Processed data for all Countries ☑️")

execute("ID")
